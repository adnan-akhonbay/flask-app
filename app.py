from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from flask import Flask, render_template, request, redirect, url_for, send_from_directory
from werkzeug.utils import secure_filename
from selenium.webdriver.chrome.service import Service

import time
total_time = time.time()
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
from pprint import pprint
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side

import os

import pandas
from fileinput import filename

# "C:\Users\adnan\AppData\Local\Packages\PythonSoftwareFoundation.Python.3.11_qbz5n2kfra8p0\LocalCache\local-packages\Python311\Scripts\virtualenv.exe"
# https://stackoverflow.com/questions/51119495/how-to-setup-environment-variables-for-flask-run-on-windows
# PS > python3 -m flask run
# https://www.geeksforgeeks.org/upload-and-read-excel-file-in-flask/
# Upload excel file in flask and display using pandas as html

# https://flask.palletsprojects.com/en/2.3.x/patterns/fileuploads/
# flask uploading file

app = Flask(__name__, template_folder='Template',static_folder='static')
UPLOAD_FOLDER = r'C:\Users\adnan\OneDrive\Desktop\project\udawi-flask-webinterface\static'
ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# xlpath = r'C:\Users\adnan\OneDrive\Desktop\project\excel\municipality_card_query.xlsx' # xl file path and name
# wb_obj = openpyxl.load_workbook(xlpath) # create a workbook object
# # sheet_obj = wb_obj['Claint'] # choose Sheet by name
# sheet_obj = wb_obj['Sheet1'] # choose Sheet by index
# m_col = sheet_obj.max_column # total number of columns
# m_row = sheet_obj.max_row # total number of rows



def whichDriver(choose):
    print('Type 1 for Firefox. 2 for Chrome. 3 for Firefox headless. 4 for Chrome headless.')
    # choose = input()
    choose =str(choose)
    if choose == '1':
        print('Launching Firefox')
        # Using Firefox
        driver = webdriver.Firefox()
        driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
        return driver
    elif choose == '2':
        print('Launching Chrome')
        # Using Chrome
        path_to_chromedriver = r"C:/Users/adnan/OneDrive/Desktop/project/chromedriver_win32/chromedriver.exe"
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        driver = webdriver.Chrome()
        # driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
        return driver
    elif choose == '3':
        # Using firefox headless
        path_to_chromedriver = r"C:/Users/adnan/OneDrive/Desktop/project/chromedriver_win32/chromedriver.exe"
        fireFoxOptions = webdriver.fireFoxOptions()
        
        fireFoxOptions.add_argument('--headless')
        driver = webdriver.Firefox(options=fireFoxOptions)
        driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
        return driver
    elif choose == '4':
        print('Launching Chrome headless')
        # Using Chrome headless
        path_to_chromedriver = r"C:/Users/adnan/OneDrive/Desktop/project/chromedriver_win32/chromedriver.exe"
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        driver = webdriver.Chrome(options=options)
        driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
        return driver

def balady(filename, driver ):
    print(filename)
    # Open an xl workbook
    # xl file path
    xlpath = r'C:\Users\adnan\OneDrive\Desktop\project\udawi-flask-webinterface\static\uploaded_xl_file'
    xlpath = os.path.join(xlpath,filename)
    wb_obj = openpyxl.load_workbook(xlpath) # create a workbook object
    # choose Sheet by name
    sheet_obj = wb_obj['Sheet1']

    # total number of columns
    m_col = sheet_obj.max_column
    # print total number of rows
    m_row = sheet_obj.max_row
    # count matches
    matchCount = 0
    idType = ''
    for i in range(2, m_row + 1):
        state = False
        # Counting time
        start_time = time.time()

        # getting cells value from 
        # the first column or ids column
        # starting from row 2
        # from xl sheet
        cell_obj = sheet_obj.cell(row = i, column = 2)
        # reading the cell value
        id = cell_obj.value

        # deciding national id, iqamah, or border id
        if str(cell_obj.value)[0] == '1':
            idType = 'national'
        elif str(cell_obj.value)[0] == '2':
            idType = 'iqamah'
        # elif str(cell_obj.value)[0] == '3':
        else:
            idType = 'border'
        
        # Calling func method enter id and idType
        time.sleep(4)
        '''
            This function will get the url, fill-in the id field and ref number field, and submit the form.
        '''
        # get url
        driver.get('https://apps.balady.gov.sa/Eservices/health/Inquiries/Search')
        driver.maximize_window()

        # find id input field
        idnumber = driver.find_element(By.XPATH, "//input[@id='IdentityNumber']")
        time.sleep(2)
        # fill-in the id field
        idnumber.send_keys(id)
        time.sleep(1)

        # find the type-of-id field
        nationalid = driver.find_element(By.ID, 'select2-IdentityType-container')
        time.sleep(3)
        # open the dropdown menu
        nationalid.click()
        time.sleep(1)
        
        # find type of id
        dropdownMenu = driver.find_element(By.ID, 'IdentityType')
        # choose an option from the select element
        select = Select(dropdownMenu)
        # choose the appropriate id type name
        if idType == 'national':
            select.select_by_index(1)
        elif idType == 'iqamah':
            select.select_by_index(2)
        elif idType == 'border':
            select.select_by_index(3)

        # # find the search button
        # submit = driver.find_element(By.XPATH, "//div[@id='Buttons']/input[1]")
        # # click the search button
        # submit.click()



        # waits
        wait = WebDriverWait(driver, 4)
        # element = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@id='Buttons']/input[1]")))
        element = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@id='Buttons']/input[1]")))
        # element.click()

        # actions
        actions = ActionChains(driver)
        time.sleep(1)
        actions.move_to_element(element).pause(1).click().pause(2).perform()
        # butts = driver.find_elements(By.ID, 'Buttons')
        # print(len(butts))
        # print(driver.execute_script("document.getElementById('Buttons').click();"))
        
        '''
                Search button clicked
            Traceback (most recent call last):
            File "/Users/AdnanAkhonbay/Desktop/Projects/MIZA/PyFiles/balady.py", line 118, in <module>
                func(id, idType)
            File "/Users/AdnanAkhonbay/Desktop/Projects/MIZA/PyFiles/balady.py", line 77, in func
                print(submit.click())
                    ^^^^^^^^^^^^^^
            File "/usr/local/lib/python3.11/site-packages/selenium/webdriver/remote/webelement.py", line 94, in click
                self._execute(Command.CLICK_ELEMENT)
            File "/usr/local/lib/python3.11/site-packages/selenium/webdriver/remote/webelement.py", line 395, in _execute
                return self._parent.execute(command, params)
                    ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^
            File "/usr/local/lib/python3.11/site-packages/selenium/webdriver/remote/webdriver.py", line 346, in execute
                self.error_handler.check_response(response)
            File "/usr/local/lib/python3.11/site-packages/selenium/webdriver/remote/errorhandler.py", line 245, in check_response
                raise exception_class(message, screen, stacktrace)
            selenium.common.exceptions.ElementClickInterceptedException: Message: element click intercepted: Element <input type="button" onclick="getCertDetails()" class="btn btn-primary" value="بحث"> is not clickable at point (1112, 434). Other element would receive the click: <div class="preloader checking" id="preloader-logo" style="display: block;">...</div>

        '''

        # end of fun
        time.sleep(3)

        try:

            # Reading error message
            alertMessage = driver.find_element(By.CLASS_NAME, "alert-danger")
            # writing error message to xl sheet
            sheet_obj.cell(row = i , column =  16).value = alertMessage.text
            # saving xl sheet
            wb_obj.save(xlpath)

            # Printing process time
            print('%s - RECORD NOT FOUND. \n Process time is: ' %(i-1), round(time.time() - start_time,1))
            state = True
            time.sleep(2)

        except:
            time.sleep(2)
            # Finding table headers
            # and assigning to variable
            th = driver.find_elements(By.TAG_NAME,'th')
            td = driver.find_elements(By.TAG_NAME,'td')

            time.sleep(2)
            # Find municipality name
            # id="MunicapilityName"
            MunicapilityName = driver.find_element(By.ID, 'MunicapilityName').get_attribute('value')
            # write municipality name to xl file
            sheet_obj.cell(row = i , column = len(td) + 4).value = MunicapilityName
            
            # Find sub-municipality name
            # id="subMunicapilityName"
            subMunicapilityName = driver.find_element(By.ID, 'subMunicapilityName').get_attribute('value')
            # writing the subMunicipalityName to xl sheet
            sheet_obj.cell(row = i , column = len(td) + 5).value = subMunicapilityName
            
            sheet_obj.cell(row = 1 , column = 4).value = 'الاسم'
            sheet_obj.cell(row = 1 , column = 5).value = 'رقم الهوية'
            sheet_obj.cell(row = 1 , column = 6).value = 'رقم الطلب'
            sheet_obj.cell(row = 1 , column = 7).value = 'رقم الفاتورة'
            sheet_obj.cell(row = 1 , column = 8).value = 'تاريخ الفاتورة'
            sheet_obj.cell(row = 1 , column = 9).value = 'حالة الفاتورة'
            sheet_obj.cell(row = 1 , column = 10).value = 'اسم المركز الطبي'
            sheet_obj.cell(row = 1 , column = 11).value = 'رقم الكشف الطبي'
            sheet_obj.cell(row = 1 , column = 12).value = 'رقم السهادة الصحية'
            sheet_obj.cell(row = 1 , column = 13).value = 'نوع الشهادة الصحية'
            sheet_obj.cell(row = 1 , column = 14).value = 'تكلفة الرسوم الاجمالية'
            sheet_obj.cell(row = 1 , column = 15).value = 'حالة الطلب'
            sheet_obj.cell(row = 1 , column = 16).value = "الاماتة"
            sheet_obj.cell(row = 1 , column = 17).value = subMunicapilityName = "البلدية"

            # Looping through headers and data
            for d in range(len(td)):
                # Openpyxl Write Data to Cell
                sheet_obj.cell(row = i , column = d + 4).value = td[d].text
                time.sleep(1)
            
            
            # Save workbook
            wb_obj.save(xlpath)
            print(i-1,'RECORD FOUND >>>> process time is: ', round(time.time() - start_time,1))
            print(i-1,'-',td[1].text,',,,',id)
            state = True
            matchCount = matchCount + 1
        finally:
            if state == False:
                print('Try searching again. >>>>> ',i-1,'-',id,' interrupted. finally.\n')
                sheet_obj.cell(row = i , column = 2).value = "Try searching again. >>>>> interrupted. For id: " + str(id)
                wb_obj.save(xlpath)
                continue
            if state == True:
                print(i-1,' - Process Finalized successfully,,,' , id, '.\n')
                continue

    software_time = time.time() - total_time
    if software_time > 60:
        software_time = software_time / 60
    print('>>>> Total time is: %s, \n ...and %s records matched.' %(round(software_time, 1),matchCount), '\n')

def fillFormNCallback(studentRowIndex, id, ref,xlpath,wb_obj,sheet_obj,m_col,m_row,driver):
    driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
    time.sleep(3)
    # Switch frame by id
    driver.switch_to.frame('subwindow')
    inputs = driver.find_elements(By.TAG_NAME, 'input')
    time.sleep(3)
    # Input fields
    inputs[17].send_keys(id)
    inputs[18].send_keys(ref)
    time.sleep(6)
    inputs[19].click()  # Submitting input field
    time.sleep(9)

def screenshotMainPage(studentRowIndex,id,ref,xlpath,wb_obj,sheet_obj,m_col,m_row,driver,filename):
    

    driver.switch_to.default_content()
    # driver.switch_to.frame('iframeSearch')
    # driver.switch_to.frame('subwindow')
    driver.set_window_size(1920, 2200)
    body = driver.find_element(By.TAG_NAME, 'body') # getting the receipt body
    html = driver.find_element(By.TAG_NAME, 'html') # getting the receipt body
    form = driver.find_element(By.TAG_NAME, 'form') # getting the receipt body
    studentName = sheet_obj.cell(row = studentRowIndex , column = 1).value
    studentId = sheet_obj.cell(row = studentRowIndex , column = 2).value
    # studentId = sheet_obj.cell(row = studentRowIndex , column = 4).value
    studentRef = sheet_obj.cell(row = studentRowIndex , column = 3).value
    
    # screenShotsPath = r"C:\Users\adnan\OneDrive\Desktop\project\screenshots\arkkan - "+studentName+" - "+str(studentId)+" - "+str(studentRef)+" -"+".png"
    screenShotsPath = os.path.join(UPLOAD_FOLDER,'screenshots',f"arkkan-{studentId}.png")
    print(screenShotsPath)
    # body.screenshot(screenShotsPath) # taking a screenshot
    html.screenshot(screenShotsPath) # taking a screenshot
    # form.screenshot(screenShotsPath) # taking a screenshot

    file_name = ''
    file_dir= screenShotsPath


    # studentId = sheet_obj.cell(row = studentRowIndex , column = 4).value
    text1 = sheet_obj.cell(row = studentRowIndex , column = 5).value
    text2 = sheet_obj.cell(row = studentRowIndex , column = 6).value
    text3 = sheet_obj.cell(row = studentRowIndex , column = 7).value
    text4 = sheet_obj.cell(row = studentRowIndex , column = 8).value
    text5 = sheet_obj.cell(row = studentRowIndex , column = 9).value
    text6 = sheet_obj.cell(row = studentRowIndex , column = 10).value
    text7 = sheet_obj.cell(row = studentRowIndex , column = 11).value
    text8 = sheet_obj.cell(row = studentRowIndex , column = 12).value
    text9 = sheet_obj.cell(row = studentRowIndex , column = 13).value
    text10 = sheet_obj.cell(row = studentRowIndex , column = 14).value
    text11 = sheet_obj.cell(row = studentRowIndex , column = 15).value
    text12 = sheet_obj.cell(row = studentRowIndex , column = 16).value
    text13 = sheet_obj.cell(row = studentRowIndex , column = 17).value

    time.sleep(4)
    if text4:
        if len(text4) == 10:
            datetime_string = text4
            datetime_string = text4
            format_string = "%Y/%m/%d"
            expiration_date = datetime.strptime(datetime_string, format_string).date() # Convert string to date using strptime
            today = datetime.today().date()
            days_difference = expiration_date -  today # Difference between dates in days
            driver.get('http://127.0.0.1:5000/arkkan-image/'+str(studentRowIndex))
            time.sleep(4)
            driver.set_window_size(1320, 2700)
            saveLocation = os.path.join(UPLOAD_FOLDER,'screenshots','arkkan-w-balady-information-'+str(studentId)+'.png')
            body = driver.find_element(By.TAG_NAME,'body')
            body.screenshot(saveLocation)
        else:
            pass
    else:
        pass
     
def screenShotBalady(studentRowIndex, id, ref,xlpath, wb_obj,sheet_obj,m_col,m_row,driver):
    studentName = sheet_obj.cell(row = studentRowIndex , column = 1).value
    studentId = sheet_obj.cell(row = studentRowIndex , column = 2).value
    studentRef = sheet_obj.cell(row = studentRowIndex , column = 3).value
    idType = str(sheet_obj.cell(row = studentRowIndex, column = 2).value)[0]
    
    if idType == '1':
        idType = 'national'
    elif idType == '2':
        idType = 'iqamah'
    else:
        idType = 'border'

    driver.get('https://apps.balady.gov.sa/Eservices/health/Inquiries/Search')
    driver.maximize_window()
    idnumber = driver.find_element(By.XPATH, "//input[@id='IdentityNumber']")
    time.sleep(2)
    idnumber.send_keys(studentId)
    time.sleep(1)
    nationalid = driver.find_element(By.ID, 'select2-IdentityType-container')
    time.sleep(3)
    nationalid.click()
    time.sleep(1)
    dropdownMenu = driver.find_element(By.ID, 'IdentityType')
    select = Select(dropdownMenu)
    if idType == 'national':
        select.select_by_index(1)
    elif idType == 'iqamah':
        select.select_by_index(2)
    elif idType == 'border':
        select.select_by_index(3)

    wait = WebDriverWait(driver, 4)  # waits
    element = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@id='Buttons']/input[1]")))
    actions = ActionChains(driver) # actions
    time.sleep(1)
    actions.move_to_element(element).pause(1).click().pause(2).perform()
    driver.set_window_size(1520, 1600)
    body = driver.find_element(By.TAG_NAME,'body')
    html = driver.find_element(By.TAG_NAME,'html')
    # screenShotsPath = r"C:\Users\adnan\OneDrive\Desktop\project\screenshots\balady - "+studentName+" - "+str(studentId)+" - "+" -"+".png"
    screenShotsPath = os.path.join(UPLOAD_FOLDER,'screenshots',f'balady-{studentId}.png')
    print(screenShotsPath)
    # body.screenshot(screenShotsPath) # taking a screenshot
    html.screenshot(screenShotsPath) # taking a screenshot

    state = False
    try:

        # Reading error message
        alertMessage = driver.find_element(By.CLASS_NAME, "alert-danger")
        # writing error message to xl sheet
        sheet_obj.cell(row = studentRowIndex , column =  18).value = alertMessage.text
        # saving xl sheet
        wb_obj.save(xlpath)

        # Printing process time
        print('%s - RECORD NOT FOUND. \n Process time is: ' %(studentRowIndex-1), "round(time.time() - start_time,1)")
        state = False
        time.sleep(2)

    except:
        time.sleep(2)
        # Finding table headers
        # and assigning to variable
        th = driver.find_elements(By.TAG_NAME,'th')
        td = driver.find_elements(By.TAG_NAME,'td')
        time.sleep(2)
        # Find municipality name
        # id="MunicapilityName"
        MunicapilityName = driver.find_element(By.ID, 'MunicapilityName').get_attribute('value')
        # write municipality name to xl file
        sheet_obj.cell(row = studentRowIndex , column = len(td) + 2).value = MunicapilityName
        
        # Find sub-municipality name
        # id="subMunicapilityName"
        subMunicapilityName = driver.find_element(By.ID, 'subMunicapilityName').get_attribute('value')
        # writing the subMunicipalityName to xl sheet
        sheet_obj.cell(row = studentRowIndex , column = len(td) + 3).value = subMunicapilityName

        sheet_obj.cell(row = 1, column = 16).value = "الامانة"
        sheet_obj.cell(row = 1, column = 17).value = "البلدية"
        sheet_obj.cell(row = studentRowIndex , column = 16).value = driver.find_element(By.ID, 'MunicapilityName').get_attribute('value')
        sheet_obj.cell(row = studentRowIndex , column = 17).value = driver.find_element(By.ID, 'subMunicapilityName').get_attribute('value')

        for i,header in enumerate(th):
            sheet_obj.cell(row = 1, column = i + 4).value = header.text        

        # Looping through headers and data
        for d in range(len(td)):
            # Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = d + 4).value = td[d].text

        # Save workbook
        wb_obj.save(xlpath)
        print(studentRowIndex-1,'RECORD FOUND >>>> process time is: ', "round(time.time() - start_time,1)")
        print(studentRowIndex-1,'-',td[1].text,',,,',id)
        state = True
    finally:
        if state == False:
            print('Try searching again. >>>>> ',studentRowIndex-1,'-',id,' interrupted. finally.\n')
            sheet_obj.cell(row = studentRowIndex , column = 4).value = "Try searching again. >>>>> interrupted. For id: " + str(id)
            wb_obj.save(xlpath)
            # continue
        if state == True:
            print(studentRowIndex-1,' - Process Finalized successfully,,,' , id, '.\n')
            # continue

def getIdsNRefs(callback1, callback2, callback3,xlpath,wb_obj,sheet_obj,m_col,m_row,driver,filename):
    '''
    This is the fild url to be called
    getIdsNRefs(func<desired end result>, query<button index>,
                callback1<fillFormNCallback>, callback2<click button and open frame>)
    '''
    for i in range(2, m_row + 1): # loop through column 1 starting at row 2
        state = False  
        start_time = time.time() # Counting time for performance testing
        id = sheet_obj.cell(row = i, column = 2).value # student id number finding and reading the cell value 
        ref = sheet_obj.cell(row = i, column = 3).value # student ref number
        callback3(i, id, ref,xlpath,wb_obj,sheet_obj,m_col,m_row,driver) # screenShotBalady
        callback1(i,id,ref,xlpath,wb_obj,sheet_obj,m_col,m_row,driver) # fillFormCallback(id,ref,'65',getFrameUrl,attendance)
        callback2(i,id,ref,xlpath,wb_obj,sheet_obj,m_col,m_row,driver,filename) # screenShotMainPage
        # func(i) # attendance(studentRowIndex)



def fillFormNSubmit(id, ref,driver):
    driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
    driver.maximize_window()
    time.sleep(3)
    # Switch frame by id
    driver.switch_to.frame('subwindow')
    inputs = driver.find_elements(By.TAG_NAME, 'input')
    time.sleep(3)
    # Input fields
    inputs[17].send_keys(id)
    inputs[18].send_keys(ref)
    time.sleep(6)
    inputs[19].click()  # Submitting input field
    time.sleep(5)

def openPopup(query,driver):
    inputs = driver.find_elements(By.TAG_NAME, 'input', )   # Getting the different button (input tags)
    if query == '65':
        try:
            e = driver.find_element(By.XPATH,'//*[@id="ctl00_Courses_Students_GridView1"]/tbody/tr[2]/td[9]/input')
            e.click()
            # inputs[65].click()   # Open attendance page
        except:
            print('Attendance Clicking the button went wrong >>> something went wrong')
            '''
                >>> Make sure you append to a collection and run again
            '''
        else:
            pass
        finally:
            '''
            
            >>> Make a code here to write error to excel sheet. Student record needed. Row number etc.
            >>> Then make sure you run the query again.

            '''
            pass

    elif query == '66':
        try:
            e = driver.find_element(By.XPATH,'//*[@id="ctl00_Courses_Students_GridView1"]/tbody/tr[2]/td[10]/input')
            e.click()
            # inputs[66].click()   # Open attendance page
        except:
            print('Exams Clicking the button went wrong >>> something went wrong')
            '''
                >>> Make sure you append to a collection and run again
            '''
        else:
            pass
        finally:
            '''
            
            >>> Make a code here to write error to excel sheet. Student record needed. Row number etc.
            >>> Then make sure you run the query again.

            '''
            pass

def attendance(studentRowIndex, xlpath, wb_obj, sheet_obj, m_col, m_row, driver):
    studentID = sheet_obj.cell(row = studentRowIndex , column = 1)
    '''
    # Receipt page. Input[65]
    Only perform pure webscraping in this page.
    Previous method already opened a new page with the following url
    driver.get('https://arkkanapp.net/Arkan/frm8158_Students.aspx?website=1&ScrId=8158&Studentsid=056fd5b7-d159-4639-861d-0f694fc03d3d&Courses_Studentsid=5f5c880b-61a5-4f63-874a-56e4ff18186e') #URL to classes registration receipt
    # Opened attendance page
    '''
    try:    
        driver.switch_to.default_content()
        driver.switch_to.frame('iframeSearch')
    except Exception as e:
        
        print('Attendance Switching to frame >>> something went wrong')
        print(e)
        sheet_obj.cell(row = studentRowIndex , column = 4).value = 'Something went wrong.'
        '''
            >>> Append to a list or dictionary and run again or pass to finally
        '''

    else:
        tr = driver.find_elements(By.XPATH, "//table[@id='ctl00_attendance_Sudents_GridView1']/tbody/tr")
        th = driver.find_elements(By.XPATH, "//table[@id='ctl00_attendance_Sudents_GridView1']/tbody/tr/th")
        td = driver.find_elements(By.XPATH, "//table[@id='ctl00_attendance_Sudents_GridView1']/tbody/tr/td")
        # color coming from: >>> https://www.computerhope.com/cgi-bin/htmlcolor.pl?c=FFAE42
        light_yellow_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
        sun_yellow_fill = PatternFill(start_color='FFE87C', end_color='FFE87C', fill_type='solid')
        neon_orange_fill = PatternFill(start_color='FF6700', end_color='FF6700', fill_type='solid')
        light_grey = PatternFill(start_color='E0DFDF', end_color='E0DFDF', fill_type='solid')
        grey = PatternFill(start_color='BAB9B8', end_color='BAB9B8', fill_type='solid')
        light_blue = PatternFill(start_color='C2E0FB', end_color='C2E0FB', fill_type='solid')
        red_fill = PatternFill(start_color='f2bb07', end_color='f2bb07', fill_type='solid')
        red_font = "f21707"


        sheet_obj.cell(row = 1 , column = 1).value = "الاسم" # Openpyxl Write Data to Cell
        sheet_obj.cell(row = 1 , column = 2).value = "الهوية" # Openpyxl Write Data to Cell
        sheet_obj.cell(row = 1 , column = 3).value = "الرقم المرجعي" # Openpyxl Write Data to Cell
        sheet_obj.cell(row = 1 , column = 1).fill = grey
        sheet_obj.cell(row = 1 , column = 2).fill = grey
        sheet_obj.cell(row = 1 , column = 3).fill = grey
        sheet_obj.cell(row = 1 , column = 3).fill = grey

        # sheet_obj.cell(row = 1 , column = 1).border = Border(top="double", left="double", right="double", bottom="double")
        # sheet_obj.cell(row = 1 , column = 2).border = Border(top="double", left="double", right="double", bottom="double")
        # sheet_obj.cell(row = 1 , column = 3).border = Border(top="double", left="double", right="double", bottom="double")


        sheet_obj.cell(row = 1 , column = 1).font = Font(bold=True)
        sheet_obj.cell(row = 1 , column = 2).font = Font(bold=True)
        sheet_obj.cell(row = 1 , column = 3).font = Font(bold=True)
        # Day 1
        sheet_obj.cell(row = 1 , column = 4).value = "المكان" # Openpyxl Write Data to Cell
        sheet_obj.cell(row = 1 , column = 5).value = "الدورة" # Openpyxl Write Data to Cell
        
        sheet_obj.cell(row = 1 , column = 4).fill = light_yellow_fill
        sheet_obj.cell(row = 1 , column = 5).fill = light_yellow_fill

        sheet_obj.cell(row = 1 , column = 4).font = Font(bold=True)
        sheet_obj.cell(row = 1 , column = 5).font = Font(bold=True)
        
        # sheet_obj.cell(row = 1 , column = 4).border = Border(top="double", left="double", right="double", bottom="double")
        # sheet_obj.cell(row = 1 , column = 5).border = Border(top="double", left="double", right="double", bottom="double")

        sheet_obj.cell(row = 1 , column = 6).value = "اليوم الاول" # Openpyxl Write Data to Cell
        sheet_obj.cell(row = 1 , column = 7).value = "الحضور اساسي" # Openpyxl Write Data to Cell
        sheet_obj.cell(row = 1 , column = 8).value = "انصراف" # Openpyxl Write Data to Cell
        
        sheet_obj.cell(row = 1 , column = 6).fill = neon_orange_fill
        sheet_obj.cell(row = 1 , column = 7).fill = neon_orange_fill
        sheet_obj.cell(row = 1 , column = 8).fill = neon_orange_fill 

        sheet_obj.cell(row = 1 , column = 6).font = Font(bold=True)
        sheet_obj.cell(row = 1 , column = 7).font = Font(bold=True)
        sheet_obj.cell(row = 1 , column = 8).font = Font(bold=True) 

        # sheet_obj.cell(row = 1 , column = 6).border = Border(top="double", left="double", right="double", bottom="double")
        # sheet_obj.cell(row = 1 , column = 7).border = Border(top="double", left="double", right="double", bottom="double")
        # sheet_obj.cell(row = 1 , column = 8).border = Border(top="double", left="double", right="double", bottom="double") 


        # Day 2
        sheet_obj.cell(row = 1 , column = 9).value = "اليوم الثاني" # Openpyxl Write Data to Cell
        sheet_obj.cell(row = 1 , column = 10).value = "الحضور اساسي" # Openpyxl Write Data to Cell
        sheet_obj.cell(row = 1 , column = 11).value = "انصراف" # Openpyxl Write Data to Cell

        sheet_obj.cell(row = 1 , column = 9).fill = sun_yellow_fill
        sheet_obj.cell(row = 1 , column = 10).fill = sun_yellow_fill
        sheet_obj.cell(row = 1 , column = 11).fill = sun_yellow_fill

        sheet_obj.cell(row = 1 , column = 9).font = Font(bold=True)
        sheet_obj.cell(row = 1 , column = 10).font = Font(bold=True)
        sheet_obj.cell(row = 1 , column = 11).font = Font(bold=True)

        # sheet_obj.cell(row = 1 , column = 9).border = Border(top="double", left="double", right="double", bottom="double")
        # sheet_obj.cell(row = 1 , column = 10).border = Border(top="double", left="double", right="double", bottom="double")
        # sheet_obj.cell(row = 1 , column = 11).border = Border(top="double", left="double", right="double", bottom="double")

         # Day 3
        sheet_obj.cell(row = 1 , column = 12).value = "اليوم الثالث" # Openpyxl Write Data to Cell
        sheet_obj.cell(row = 1 , column = 13).value = "الحضور اساسي" # Openpyxl Write Data to Cell
        sheet_obj.cell(row = 1 , column = 14).value = "انصراف" # Openpyxl Write Data to Cell

        sheet_obj.cell(row = 1 , column = 12).fill = light_blue
        sheet_obj.cell(row = 1 , column = 13).fill = light_blue
        sheet_obj.cell(row = 1 , column = 14).fill = light_blue

        sheet_obj.cell(row = 1 , column = 12).font = Font(bold=True)
        sheet_obj.cell(row = 1 , column = 13).font = Font(bold=True)
        sheet_obj.cell(row = 1 , column = 14).font = Font(bold=True)

        # sheet_obj.cell(row = 1 , column = 12).border = Border(top="double", left="double", right="double", bottom="double")
        # sheet_obj.cell(row = 1 , column = 13).border = Border(top="double", left="double", right="double", bottom="double")
        # sheet_obj.cell(row = 1 , column = 14).border = Border(top="double", left="double", right="double", bottom="double")


        # day1.fill = light_yellow_fill
        
        # day2 = sheet_obj.cell(row = 1 , column = h+6 + 4)
        # day2.value = th[h].text # Openpyxl Write Data to Cell
        # day2.fill = sun_yellow_fill

        # day3 = sheet_obj.cell(row = 1 , column = h+11 + 4)
        # day3.value = th[h].text # Openpyxl Write Data to Cell
        # day3.fill = neon_orange_fill

        studentName = sheet_obj.cell(row = studentRowIndex , column = 1).value
        screenShotsPath = r"C:\Users\adnan\OneDrive\Desktop\project\excel\screenshots"
        screenShotsPath = os.path.join(screenShotsPath,str(studentName)+".png" )
        driver.find_element(By.TAG_NAME,'body').screenshot(screenShotsPath)

        if len(td) == 18: # three days
            sheet_obj.cell(row = studentRowIndex , column = 4).value = td[13].text # Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = 5 ).value = td[14].text # Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = 6 ).value = td[15].text # Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = 9 ).value = td[9].text # Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = 12 ).value = td[3].text # Openpyxl Write Data to Cell

            td4 = td[4]
            xl13 = sheet_obj.cell(row = studentRowIndex , column = 13 )
            xl13.value = td[4].text # Openpyxl Write Data to Cell

            td5 = td[5]
            xl14 = sheet_obj.cell(row = studentRowIndex , column = 14 )
            xl14.value = td5.text # Openpyxl Write Data to Cell

            td10 = td[10]
            xl10 = sheet_obj.cell(row = studentRowIndex , column = 10 )
            xl10.value = td10.text # Openpyxl Write Data to Cell  

            td11 = td[11]
            xl11 = sheet_obj.cell(row = studentRowIndex , column = 11 )
            xl11.value = td11.text # Openpyxl Write Data to Cell

            td16 = td[16]
            xl7 = sheet_obj.cell(row = studentRowIndex , column = 7 )
            xl7.value = td16.text # Openpyxl Write Data to Cell          
            
            td17 = td[17]
            xl8 = sheet_obj.cell(row = studentRowIndex , column = 8 )
            xl8.value = td17.text


            if td16.text == "غ":
                xl7.font = Font(red_font)
                xl7.fill = red_fill
                
            if td17.text == "غ":
                xl8.font = Font(color=red_font)
                xl8.fill = red_fill
            if td10.text == "غ":
                xl10.font = Font(color=red_font)
                xl10.fill = red_fill
            if td11.text == "غ":
                xl11.font = Font(color=red_font)
                xl11.fill = red_fill
            if td4.text == "غ":
                xl13.font = Font(color=red_font)
                xl13.fill = red_fill
            if td5.text == "غ":
                xl14.font = Font(color=red_font)
                xl14.fill = red_fill


            wb_obj.save(xlpath)
        elif len(td) == 12:
            sheet_obj.cell(row = studentRowIndex , column = 4 ).value = td[7].text # Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = 5 ).value = td[8].text # Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = 6 ).value = td[9].text # Openpyxl Write Data to Cell

            td10 = td[10]
            xl7 = sheet_obj.cell(row = studentRowIndex , column = 7 )
            xl7.value = td10.text # Openpyxl Write Data to Cell          
            td11 = td[11]
            xl8 = sheet_obj.cell(row = studentRowIndex , column = 8 )
            xl8.value = td11.text # Openpyxl Write Data to Cell
            
            sheet_obj.cell(row = studentRowIndex , column = 9 ).value = td[3].text # Openpyxl Write Data to Cell


            td4 = td[4]
            xl10 = sheet_obj.cell(row = studentRowIndex , column = 10)
            xl10.value = td4.text # Openpyxl Write Data to Cell          
            td5 = td[5]
            xl11 = sheet_obj.cell(row = studentRowIndex , column = 11 )
            xl11.value = td5.text # Openpyxl Write Data to Cell

            if td4.text == "غ":
                xl10.fill = red_fill
                xl10.font = Font(color=red_font)
                
            if td5.text == "غ":
                xl11.fill = red_fill
                xl11.font = Font(color=red_font)

            if td10.text == "غ":
                xl7.fill = red_fill
                xl7.font = Font(color=red_font)
                
            if td11.text == "غ":
                xl8.fill = red_fill
                xl8.font = Font(color=red_font)

            wb_obj.save(xlpath) 

        elif len(td) == 6:

            sheet_obj.cell(row = studentRowIndex , column = 4 ).value = td[1].text # Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = 5 ).value = td[2].text # Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = 6 ).value = td[3].text # Openpyxl Write Data to Cell

            td4 = td[4]
            xl7 = sheet_obj.cell(row = studentRowIndex , column = 7 )
            xl7.value = td[4].text # Openpyxl Write Data to Cell          
            td5 = td[5]
            xl8 = sheet_obj.cell(row = studentRowIndex , column = 8 )
            xl8.value = td[5].text # Openpyxl Write Data to Cell


            if td4.text == "غ":
                xl7.fill = red_fill
                xl7.font = Font(color=red_font)
                
            if td5.text == "غ":
                xl8.fill = red_fill
                xl8.font = Font(color=red_font)
            
            wb_obj.save(xlpath)
        
    finally:
        screenShotsPath = os.path.join(UPLOAD_FOLDER,'screenshots',f"attendance-{sheet_obj.cell(row = studentRowIndex , column = 1).value}.png")
        print(screenShotsPath)
        body = driver.find_element(By.TAG_NAME, 'body')
        body.screenshot(screenShotsPath) # taking a screenshot
        # html.screenshot(screenShotsPath) # taking a screenshot
        pass
        '''


        >>> Write code to rerun process.
        
        '''
    # time.sleep(3000000) # long sleep for debugging puroses

def exams(studentRowIndex, xlpath,wb_obj, sheet_obj, m_col, m_row,driver):
    red_font = "f21707"
    '''
    # Receipt page. Input[66]
    Only perform pure webscraping in this page.
    Previous method already opened a new page with the following url
    driver.get('https://arkkanapp.net/Arkan/frm8159_Students.aspx?website=1&ScrId=8159&Studentsid=056fd5b7-d159-4639-861d-0f694fc03d3d&Courses_Dates_id=8f5cca31-2f95-4c6b-a0c1-7ee8d8884a65') #URL to classes registration receipt
    # Opened exams results
    '''
    try:
        driver.switch_to.default_content()
        driver.switch_to.frame('iframeSearch')
    except:
        print('Exams Switching to frame >>> something went wrong')
        sheet_obj.cell(row = studentRowIndex , column = 4).value = 'In Exams Something went wrong.'
        
    else:
        

        table2 = driver.find_element(By.ID, "ctl00_Exam_master3_GridView1")

        # studentName = sheet_obj.cell(row = studentRowIndex , column = 1).value
        screenShotsPath = r"C:\Users\adnan\OneDrive\Desktop\project\excel\screenshots"
        screenShotsPath = os.path.join('\\exams'+str(studentRowIndex)+".png" )
        driver.find_element(By.TAG_NAME,'body').screenshot(screenShotsPath)

        tr = driver.find_elements(By.XPATH, "//table[@id='ctl00_Exam_master2_GridView1']/tbody/tr")
        th = driver.find_elements(By.XPATH, "//table[@id='ctl00_Exam_master2_GridView1']/tbody/tr/th")
        td = driver.find_elements(By.XPATH, "//table[@id='ctl00_Exam_master2_GridView1']/tbody/tr/td")
        print('Has to be either 1 or 6: >>>>> ' ,len(td))
        if len(td) == 1:
            sheet_obj.cell(row = studentRowIndex , column = 15).value = td[0].text
        if len(td) > 1:
            yellow_lawn_green = PatternFill(start_color='87F717', end_color='87F717', fill_type='solid')
            examAttempt1 = sheet_obj.cell(row = 1 , column = 15) # الموعد
            examAttempt1.value = th[1].text # الموعد
            examAttempt1.fill = yellow_lawn_green

            sheet_obj.cell(row = 1 , column = 16).value = th[4].text # النتيجة
            sheet_obj.cell(row = 1 , column = 16).fill = yellow_lawn_green 

            sheet_obj.cell(row = 1 , column = 17).value = th[5].text # الدرجات
            sheet_obj.cell(row = 1 , column = 17).fill = yellow_lawn_green

            sheet_obj.cell(row = 1 , column = 18).value = th[6].text # الاجابات الصحيحة
            sheet_obj.cell(row = 1 , column = 18).fill = yellow_lawn_green

            sheet_obj.cell(row = studentRowIndex , column = 15).value = td[1].text # الموعد
            sheet_obj.cell(row = studentRowIndex , column = 16).value = td[4].text # النتيجة
            if td[4].text == "راسب":
                sheet_obj.cell(row = studentRowIndex , column = 16).font = Font(color=red_font)
            if td[4].text == "ناجح":
                sheet_obj.cell(row = studentRowIndex , column = 16).font = Font(color="87F717")



            sheet_obj.cell(row = studentRowIndex , column = 17).value = td[5].text # Start at column 18. Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = 18).value = td[6].text # Start at column 18. Openpyxl Write Data to Cell
            
            wb_obj.save(xlpath)


        tr = driver.find_elements(By.XPATH, "//table[@id='ctl00_Exam_master3_GridView1']/tbody/tr")
        th = driver.find_elements(By.XPATH, "//table[@id='ctl00_Exam_master3_GridView1']/tbody/tr/th")
        td = driver.find_elements(By.XPATH, "//table[@id='ctl00_Exam_master3_GridView1']/tbody/tr/td")
        if len(td) == 1:
            sheet_obj.cell(row = studentRowIndex , column = 19).value = td[0].text
        elif len(td) > 1:
            neon_orange_fill = PatternFill(start_color='FF6700', end_color='FF6700', fill_type='solid')


            sheet_obj.cell(row = 1 , column = 19).value = th[2].text # النتيجة
            sheet_obj.cell(row = 1 , column = 19).fill = neon_orange_fill

            sheet_obj.cell(row = 1 , column = 20).value = th[3].text # الدرجات
            sheet_obj.cell(row = 1 , column = 20).fill = neon_orange_fill

            sheet_obj.cell(row = 1 , column = 21).value = th[4].text #الاجابات الصحيحة
            sheet_obj.cell(row = 1 , column = 21).fill = neon_orange_fill

            sheet_obj.cell(row = studentRowIndex , column = 19).value = td[2].text # النتيجة
            if td[2].text == "راسب":
                sheet_obj.cell(row = studentRowIndex , column = 19).font = Font(color=red_font)
            if td[2].text == "ناجح":
                sheet_obj.cell(row = studentRowIndex , column = 19).font = Font(color="87F717")
            sheet_obj.cell(row = studentRowIndex , column = 20).value = td[3].text # Start at column 18. Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = 21).value = td[4].text # Start at column 18. Openpyxl Write Data to Cell
            wb_obj.save(xlpath)
        '''1

            >>> Collect repeat exam information

        1'''

        '''
            >>> Missing here writing information to excel sheet.
        '''
        # time.sleep(3000000) # long sleep for debugging puroses  
    finally:
        screenShotsPath = os.path.join(UPLOAD_FOLDER,'screenshots',f"exams-{sheet_obj.cell(row = studentRowIndex , column = 1).value}.png")
        print(screenShotsPath)
        body = driver.find_element(By.TAG_NAME, 'body')
        body.screenshot(screenShotsPath) # taking a screenshot
        # html.screenshot(screenShotsPath) # taking a screenshot
        """

            >>> Rerun the code

        """

def arkkanIdsNRefs(attendance, exams, fillFormNSubmit,openPopup,xlpath,wb_obj,sheet_obj,m_col,m_row,driver):
    funcStartingTime = time.time()
    '''
    This is the fild url to be called
    getIdsNRefs(attendance, exams,
                fillFormNSubmit,openPopu[])
    '''
    driver=whichDriver('4')
    for i in range(2, m_row + 1): # loop through column 1 starting at row 2
        state = False  
        start_time = time.time() # Counting time for performance testing
        id = sheet_obj.cell(row = i, column = 2).value # student id number finding and reading the cell value 
        ref = sheet_obj.cell(row = i, column = 3).value # student reg number
        fillFormNSubmit(id,ref,driver) # fillFormNSubmit(id,ref,'65',getFrameUrl,attendance)
        openPopup('65',driver)
        attendance(i, xlpath, wb_obj, sheet_obj, m_col, m_row, driver) # attendance(studentRowIndex)
        fillFormNSubmit(id,ref,driver) # fillFormNSubmit(id,ref,'65',getFrameUrl,attendance)
        openPopup('66', driver)
        exams(i, xlpath, wb_obj, sheet_obj, m_col, m_row, driver) # attendance(studentRowIndex)
    funcEndTime = time.time()

    print('FUNC TIME: ', funcEndTime - funcStartingTime)










@app.route("/attendance-and-exams", methods=['Get','POST']) # upload excel files
def attendance_and_exams():

    # # check if the post request has the file part
    # if 'file' not in request.files:
    #     flash('No file part')
    #     return redirect(request.url)

    file = request.files['file']
    # # If the user does not select a file, the browser submits an
    # # empty file without a filename.
    # if file.filename == '':
    #     flash('No selected file')
    #     return redirect(request.url)
    # if file and allowed_file(file.filename):

    filename = secure_filename(file.filename)
    print(filename)
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'uploaded_xl_file', filename))
    openFile = os.path.join(UPLOAD_FOLDER,'uploaded_xl_file',file.filename)
    # xlpath = r'C:\Users\adnan\OneDrive\Desktop\project\excel\municipality_card_query.xlsx' # xl file path and name
    xlpath = os.path.join(UPLOAD_FOLDER,'uploaded_xl_file',file.filename)
    wb_obj = openpyxl.load_workbook(xlpath) # create a workbook object
    # sheet_obj = wb_obj['Claint'] # choose Sheet by name
    sheet_obj = wb_obj['Sheet1'] # choose Sheet by index
    m_col = sheet_obj.max_column # total number of columns
    m_row = sheet_obj.max_row # total number of rows
    openFile = os.path.join(UPLOAD_FOLDER,'uploaded_xl_file',file.filename)
    driver = whichDriver('4')
    arkkanIdsNRefs(attendance, exams, fillFormNSubmit, openPopup, xlpath, wb_obj, sheet_obj, m_col, m_row, driver)

    data = pandas.read_excel(openFile) # Parse the data as a Pandas DataFrame type
    return data.to_html()
    # return redirect(url_for('download_file', name=filename)) # Download the file 


@app.route("/") # excel post form  # index page
def municipality_card_query():

    return render_template('index.html')

@app.route("/arkkan-images", methods=['Get','POST']) # upload excel files
def postExcelFile():


    # # check if the post request has the file part
    # if 'file' not in request.files:
    #     flash('No file part')
    #     return redirect(request.url)

    file = request.files['file']
    # # If the user does not select a file, the browser submits an
    # # empty file without a filename.
    # if file.filename == '':
    #     flash('No selected file')
    #     return redirect(request.url)
    # if file and allowed_file(file.filename):

    filename = secure_filename(file.filename)
    print(filename)
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'uploaded_xl_file', filename))
    openFile = os.path.join(UPLOAD_FOLDER,'uploaded_xl_file',file.filename)
    # xlpath = r'C:\Users\adnan\OneDrive\Desktop\project\excel\municipality_card_query.xlsx' # xl file path and name
    xlpath = os.path.join(UPLOAD_FOLDER,'uploaded_xl_file',file.filename)
    wb_obj = openpyxl.load_workbook(xlpath) # create a workbook object
    # sheet_obj = wb_obj['Claint'] # choose Sheet by name
    sheet_obj = wb_obj['Sheet1'] # choose Sheet by index
    m_col = sheet_obj.max_column # total number of columns
    m_row = sheet_obj.max_row # total number of rows
    openFile = os.path.join(UPLOAD_FOLDER,'uploaded_xl_file',file.filename)
    driver = whichDriver('4')
    '''
    ///////////
    '''
    getIdsNRefs(fillFormNCallback, screenshotMainPage, screenShotBalady,xlpath,wb_obj,sheet_obj,m_col,m_row,driver,filename)
    data = pandas.read_excel(openFile) # Parse the data as a Pandas DataFrame type
    return data.to_html()
    # return redirect(url_for('download_file', name=filename)) # Download the file     

@app.route("/balady-inquery", methods=['Get','POST']) # upload excel files
def postBaladyExcelFile():


    # # check if the post request has the file part
    # if 'file' not in request.files:
    #     flash('No file part')
    #     return redirect(request.url)

    file = request.files['file']
    # # If the user does not select a file, the browser submits an
    # # empty file without a filename.
    # if file.filename == '':
    #     flash('No selected file')
    #     return redirect(request.url)
    # if file and allowed_file(file.filename):

    filename = secure_filename(file.filename)
    file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'uploaded_xl_file', filename))

    # xlpath = r'C:\Users\adnan\OneDrive\Desktop\project\excel\municipality_card_query.xlsx' # xl file path and name
    xlpath = os.path.join(UPLOAD_FOLDER,'uploaded_xl_file',filename)
    wb_obj = openpyxl.load_workbook(xlpath) # create a workbook object
    # sheet_obj = wb_obj['Claint'] # choose Sheet by name
    sheet_obj = wb_obj['Sheet1'] # choose Sheet by index
    m_col = sheet_obj.max_column # total number of columns
    m_row = sheet_obj.max_row # total number of rows
    driver = whichDriver('4')
    balady(filename,driver)
    openFile = os.path.join(UPLOAD_FOLDER,'uploaded_xl_file',file.filename)
    data = pandas.read_excel(openFile) # Parse the data as a Pandas DataFrame type
    # return redirect(url_for('download_file', name=filename)) # Download the file     
    return data.to_html()

@app.route('/arkkan-image/<int:studentRowIndex>') # download the file
def student_record(studentRowIndex):

    xlpath = r'C:\Users\adnan\OneDrive\Desktop\project\udawi-flask-webinterface\static\uploaded_xl_file' # xl file path and name
    staticpath = r'C:\Users\adnan\OneDrive\Desktop\project\udawi-flask-webinterface\static'
    # filename = str(filename)
    filename = 'bajah_balady.xlsx'
    xlpath = os.path.join(xlpath,filename)
    studentRowIndex = int(studentRowIndex)
    wb_obj = openpyxl.load_workbook(xlpath) # create a workbook object
    # sheet_obj = wb_obj['Claint'] # choose Sheet by name
    sheet_obj = wb_obj['Sheet1'] # choose Sheet by index
    # m_col = sheet_obj.max_column # total number of columns
    # m_row = sheet_obj.max_row # total number of rows
    studentId = sheet_obj.cell(row = studentRowIndex , column = 2).value
    text0 = sheet_obj.cell(row = studentRowIndex , column = 4).value
    text1 = sheet_obj.cell(row = studentRowIndex , column = 5).value
    text2 = sheet_obj.cell(row = studentRowIndex , column = 6).value
    text3 = sheet_obj.cell(row = studentRowIndex , column = 7).value
    text4 = sheet_obj.cell(row = studentRowIndex , column = 8).value
    text5 = sheet_obj.cell(row = studentRowIndex , column = 9).value
    text6 = sheet_obj.cell(row = studentRowIndex , column = 10).value
    text7 = sheet_obj.cell(row = studentRowIndex , column = 11).value
    text8 = sheet_obj.cell(row = studentRowIndex , column = 12).value
    text9 = sheet_obj.cell(row = studentRowIndex , column = 13).value
    text10 = sheet_obj.cell(row = studentRowIndex , column = 14).value
    text11 = sheet_obj.cell(row = studentRowIndex , column = 15).value
    text12 = sheet_obj.cell(row = studentRowIndex , column = 16).value
    text13 = sheet_obj.cell(row = studentRowIndex , column = 17).value
    imagename =f'arkkan-{studentId}.png'
    imagefile = os.path.join('static',imagename)
    relativestaticeimagefile = os.path.join('static',imagename)
    staticimagefile = os.path.join(staticpath,imagename)
    today = datetime.today().date()
    if text4:
        if len(text4) == 10:
            datetime_string = text4
            datetime_string = text4
            format_string = "%Y/%m/%d"
            expiration_date = datetime.strptime(datetime_string, format_string).date() # Convert string to date using strptime
            days_difference = today - expiration_date # Difference between dates in days
            days_difference = days_difference.days
        else:
            days_difference = ''
    else:
        days_difference = ''


    content = {'today':today,'expirationdate':days_difference,'imagename':imagename,'txt0':text0,'txt1':text1,'txt2':text2,'txt3':text3,'txt4':text4,'txt5':text5,'txt6':text6,'txt7':text7,'txt8':text8,'txt9':text9,'txt10':text10,'txt11':text11, 'txt12':text12, 'txt13':text13}
    return render_template('arkkan-image.html',context=content)

@app.route('/uploads/<name>') # download the file
def download_file(name):
    return send_from_directory(app.config["UPLOAD_FOLDER"], name)

    # return 'title'

if __name__=="__main":
    app.run(debug=True)