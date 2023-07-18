# https://www.youtube.com/watch?v=Zrx8FSEo9lk
# selenium open existing chrome browser
# First cd /Applications
# Google\ Chrome.app/Contents/MacOS/Google\ Chrome --remote-debugging-port=8989 --user-data-dir="/Users/AdnanAkhonbay/Desktop/Projects/MIZA/PyFiles"
# https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/
# https://www.javatpoint.com/python-openpyxl#WriteDatatoCell
# Get the value in the input text field
# https://pythonexamples.org/python-selenium-get-value-in-input-text-field/

# https://wkhtmltopdf.org/downloads.html
# C:\Program Files\wkhtmltopdf\bin

# https://www.youtube.com/watch?v=iOpGGW__oz4
# driver.get_screenshot_as_file('main_page.png') # taking a screenshot

# https://holypython.com/python-pil-tutorial/how-to-add-text-to-images-in-python-via-pil-library/?expand_article=1
# adding text to image using PIL

import pdfkit
import time
total_time = time.time()
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime
from pprint import pprint
import openpyxl
from openpyxl.styles import PatternFill
# from pynput.keyboard import Key, Controller
# import pynput.keyboard as kb
import os
# from PIL import Image, ImageDraw, ImageFont, ImageColor

# # Creating a persisitent browser session
# PATH = "C:\Program Files (x86)\chromedriver_win32" # path to chromedriver
# s = Service(PATH)
# opt = Options()
# opt.add_experimental_option("debuggerAddress", "localhost:8989")
# driver = webdriver.Chrome(service=s,options=opt)
# driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')


'''
    >>> Open the xl workbook and choose sheet[0]
'''
xlpath = r'C:\Users\adnan\OneDrive\Desktop\project\excel\municipality_card_query.xlsx' # xl file path and name
wb_obj = openpyxl.load_workbook(xlpath) # create a workbook object
# sheet_obj = wb_obj['Claint'] # choose Sheet by name
sheet_obj = wb_obj['Sheet1'] # choose Sheet by index
m_col = sheet_obj.max_column # total number of columns
m_row = sheet_obj.max_row # total number of rows

def whichDriver():
    print('Type 1 for Firefox. 2 for Chrome. 3 for Firefox headless. 4 for Chrome headless.')
    choose = input()
    if choose == '1':
        print('Launching Firefox')
        # Using Firefox
        driver = webdriver.Firefox()
        driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
        return driver
    elif choose == '2':
        print('Launching Chrome')
        # Using Chrome
        path_to_chromedriver = r"C:/Users/adnan/OneDrive/Desktop/project/chromedriver_win32/chromedriver.exe"
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        driver = webdriver.Chrome()
        # driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
        return driver
    elif choose == '3':
        # Using firefox headless
        path_to_chromedriver = r"C:/Users/adnan/OneDrive/Desktop/project/chromedriver_win32/chromedriver.exe"
        fireFoxOptions = webdriver.fireFoxOptions()
        
        fireFoxOptions.add_argument('--headless')
        driver = webdriver.Firefox(options=fireFoxOptions)
        driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
        return driver
    elif choose == '4':
        print('Launching Chrome headless')
        # Using Chrome headless
        path_to_chromedriver = r"C:/Users/adnan/OneDrive/Desktop/project/chromedriver_win32/chromedriver.exe"
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        driver = webdriver.Chrome(options=options)
        driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
        return driver

driver = whichDriver()

def fillFormNCallback(studentRowIndex, id, ref):
    driver.get('https://arkkanapp.net/Bases/MainPage.aspx?url=98A7B2')
    time.sleep(3)
    # Switch frame by id
    driver.switch_to.frame('subwindow')
    inputs = driver.find_elements(By.TAG_NAME, 'input')
    time.sleep(3)
    # Input fields
    inputs[17].send_keys(id)
    inputs[18].send_keys(ref)
    time.sleep(6)
    inputs[19].click()  # Submitting input field
    time.sleep(9)


def confirmationReceipt():
    # inputs = driver.find_elements(By.TAG_NAME, 'input', )   # Getting the different button (input tags)
    # print('64,  الايصال.......')
    # inputs[64].click()   # Open registration receipt
    print('confirmationReceipt()')
    '''
    # Receipt page. Input[64]
    Only perform pure webscraping in this page.
    Previus method already opened a new page with the following url
    driver.get('https://arkkanapp.net/Documents/sanad-course.aspx') #URL to classes registration receipt
    # Opened registration receipt    
    '''
    driver.set_window_size(1700, 6000)
    hidensanad = driver.find_element(By.ID, 'hidensanad') # getting the receipt body
    hidensanad.screenshot('hidensanad_screenshot.png') # taking a screenshot 
    # driver.get_screenshot_as_file('hidensanad_screenshot.png') # taking a screenshot 

    '''
    # >>> Missing here naming image and and choosing destination probably need OS.
    '''         
    # time.sleep(300000) # long sleep for debugging puroses

def screenshotMainPage(studentRowIndex,id,ref):
    

    driver.switch_to.default_content()
    # driver.switch_to.frame('iframeSearch')
    # driver.switch_to.frame('subwindow')
    driver.set_window_size(1920, 2500)
    body = driver.find_element(By.TAG_NAME, 'body') # getting the receipt body
    html = driver.find_element(By.TAG_NAME, 'html') # getting the receipt body
    form = driver.find_element(By.TAG_NAME, 'form') # getting the receipt body
    studentName = sheet_obj.cell(row = studentRowIndex , column = 1).value
    studentId = sheet_obj.cell(row = studentRowIndex , column = 2).value
    studentRef = sheet_obj.cell(row = studentRowIndex , column = 3).value
    
    screenShotsPath = r"C:\Users\adnan\OneDrive\Desktop\project\screenshots\arkkan - "+studentName+" - "+str(studentId)+" - "+str(studentRef)+" -"+".png"
    screenShotsPath = os.path.join(screenShotsPath)
    print(screenShotsPath)
    # body.screenshot(screenShotsPath) # taking a screenshot
    html.screenshot(screenShotsPath) # taking a screenshot
    # form.screenshot(screenShotsPath) # taking a screenshot

    file_name = ''
    file_dir= screenShotsPath


    text0 = sheet_obj.cell(row = studentRowIndex , column = 4).value
    text1 = sheet_obj.cell(row = studentRowIndex , column = 5).value
    text2 = sheet_obj.cell(row = studentRowIndex , column = 6).value
    text3 = sheet_obj.cell(row = studentRowIndex , column = 7).value
    text4 = sheet_obj.cell(row = studentRowIndex , column = 8).value
    text5 = sheet_obj.cell(row = studentRowIndex , column = 9).value
    text6 = sheet_obj.cell(row = studentRowIndex , column = 10).value
    text7 = sheet_obj.cell(row = studentRowIndex , column = 11).value
    text8 = sheet_obj.cell(row = studentRowIndex , column = 12).value
    text9 = sheet_obj.cell(row = studentRowIndex , column = 13).value
    text10 = sheet_obj.cell(row = studentRowIndex , column = 14).value
    text11 = sheet_obj.cell(row = studentRowIndex , column = 15).value



    if len(text4) == 10:
        datetime_string = text4
        datetime_string = text4
        format_string = "%Y/%m/%d"
        expiration_date = datetime.strptime(datetime_string, format_string).date() # Convert string to date using strptime
        today = datetime.today().date()
        days_difference = expiration_date -  today # Difference between dates in days
    else:
        pass
     


def screenShotBalady(studentRowIndex, id, ref):
    studentName = sheet_obj.cell(row = studentRowIndex , column = 1).value
    studentId = sheet_obj.cell(row = studentRowIndex , column = 2).value
    studentRef = sheet_obj.cell(row = studentRowIndex , column = 3).value
    idType = str(sheet_obj.cell(row = studentRowIndex, column = 2).value)[0]
    
    if idType == '1':
        idType = 'national'
    elif idType == '2':
        idType = 'iqamah'
    else:
        idType = 'border'

    driver.get('https://apps.balady.gov.sa/Eservices/health/Inquiries/Search')
    driver.maximize_window()
    idnumber = driver.find_element(By.XPATH, "//input[@id='IdentityNumber']")
    time.sleep(2)
    idnumber.send_keys(studentId)
    time.sleep(1)
    nationalid = driver.find_element(By.ID, 'select2-IdentityType-container')
    time.sleep(3)
    nationalid.click()
    time.sleep(1)
    dropdownMenu = driver.find_element(By.ID, 'IdentityType')
    select = Select(dropdownMenu)
    if idType == 'national':
        select.select_by_index(1)
    elif idType == 'iqamah':
        select.select_by_index(2)
    elif idType == 'border':
        select.select_by_index(3)

    wait = WebDriverWait(driver, 4)  # waits
    # element = wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@id='Buttons']/input[1]")))
    element = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@id='Buttons']/input[1]")))
    # element.click()

    actions = ActionChains(driver) # actions
    time.sleep(1)
    actions.move_to_element(element).pause(1).click().pause(2).perform()


    '''
    
    >>> Try to screenshoot html tag to get the enitre page.
    
    '''
    driver.set_window_size(1520, 1600)
    body = driver.find_element(By.TAG_NAME,'body')
    html = driver.find_element(By.TAG_NAME,'html')
    screenShotsPath = r"C:\Users\adnan\OneDrive\Desktop\project\screenshots\balady - "+studentName+" - "+str(studentId)+" - "+" -"+".png"
    screenShotsPath = os.path.join(screenShotsPath)
    print(screenShotsPath)
    # body.screenshot(screenShotsPath) # taking a screenshot
    html.screenshot(screenShotsPath) # taking a screenshot

    state = False
    try:

        # Reading error message
        alertMessage = driver.find_element(By.CLASS_NAME, "alert-danger")
        # writing error message to xl sheet
        sheet_obj.cell(row = studentRowIndex , column =  18).value = alertMessage.text
        # saving xl sheet
        wb_obj.save(xlpath)

        # Printing process time
        print('%s - RECORD NOT FOUND. \n Process time is: ' %(studentRowIndex-1), "round(time.time() - start_time,1)")
        state = False
        time.sleep(2)

    except:
        time.sleep(2)
        # Finding table headers
        # and assigning to variable
        th = driver.find_elements(By.TAG_NAME,'th')
        td = driver.find_elements(By.TAG_NAME,'td')
        time.sleep(2)
        # Find municipality name
        # id="MunicapilityName"
        MunicapilityName = driver.find_element(By.ID, 'MunicapilityName').get_attribute('value')
        # write municipality name to xl file
        sheet_obj.cell(row = studentRowIndex , column = len(td) + 2).value = MunicapilityName
        
        # Find sub-municipality name
        # id="subMunicapilityName"
        subMunicapilityName = driver.find_element(By.ID, 'subMunicapilityName').get_attribute('value')
        # writing the subMunicipalityName to xl sheet
        sheet_obj.cell(row = studentRowIndex , column = len(td) + 3).value = subMunicapilityName
        
        for i,header in enumerate(th):
            sheet_obj.cell(row = 1, column = i + 4).value = header.text

        # Looping through headers and data
        for d in range(len(td)):
            # Openpyxl Write Data to Cell
            sheet_obj.cell(row = studentRowIndex , column = d + 4).value = td[d].text

        # Save workbook
        wb_obj.save(xlpath)
        print(studentRowIndex-1,'RECORD FOUND >>>> process time is: ', "round(time.time() - start_time,1)")
        print(studentRowIndex-1,'-',td[1].text,',,,',id)
        state = True
    finally:
        if state == False:
            print('Try searching again. >>>>> ',studentRowIndex-1,'-',id,' interrupted. finally.\n')
            sheet_obj.cell(row = studentRowIndex , column = 4).value = "Try searching again. >>>>> interrupted. For id: " + str(id)
            wb_obj.save(xlpath)
            # continue
        if state == True:
            print(studentRowIndex-1,' - Process Finalized successfully,,,' , id, '.\n')
            # continue

def getIdsNRefs(callback1, callback2, callback3):
    '''
    This is the fild url to be called
    getIdsNRefs(func<desired end result>, query<button index>,
                callback1<fillFormNCallback>, callback2<click button and open frame>)
    '''
    for i in range(2, m_row + 1): # loop through column 1 starting at row 2
        state = False  
        start_time = time.time() # Counting time for performance testing
        id = sheet_obj.cell(row = i, column = 2).value # student id number finding and reading the cell value 
        ref = sheet_obj.cell(row = i, column = 3).value # student ref number
        callback3(i, id, ref)
        callback1(i,id,ref) # fillFormCallback(id,ref,'65',getFrameUrl,attendance)
        callback2(i,id,ref) # click button and open frame
        # func(i) # attendance(studentRowIndex)
'''

>>> End of code xl testing

'''

getIdsNRefs(fillFormNCallback, screenshotMainPage, screenShotBalady)

# screenShotBalady(2,'2195398157','refnum')

# driver.quit()

'''
Making a pdf version
# # # Define path to wkhtmltopdf.exe
# path_to_wkhtmltopdf = r"C:/Program Files/wkhtmltopdf/bin/wkhtmltopdf.exe"
# print('preparing pdf')
# # Point pdfkit configuration to wkhtmltopdf.exe
# config = pdfkit.configuration(wkhtmltopdf = path_to_wkhtmltopdf)
# # Conver webpage to PDF
# # get_url = driver.current_url
# pdfkit.from_url(url, output_path= 'webpage.pdf', configuration=config)
# print('done preparing pdf')
# # id='hidensanad'
'''

'''
# Collect events until released
def on_press(key):
    try:
        print('alphanumeric key {0} pressed'.format(key.char))
    except AttributeError:
        print('special key {0} pressed'.format(key))

def on_release(key):
    print('{0} released'.format(
        key))
    if key == keyboard.Key.esc:
        # Stop listener
        return False
with kb.Listener(on_press=on_press) as listener:
    try:
        listener.join()
    except Exception as e:
        print('{0} was pressed'.format(e.args[0]))
'''